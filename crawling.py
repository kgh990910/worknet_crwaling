# 크롤링을 위한 모듈
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By

# 엑셀로 저장하기 위한 모듈
from openpyxl import Workbook
from openpyxl.styles import Border, Side
from openpyxl.styles.fonts import Font
import datetime

# 에러 해결을 위한 모듈[AttributeError: module 'collections' has no attribute 'Callable']
import collections
if not hasattr(collections, 'Callable'):
    collections.Callable = collections.abc.Callable

# 사용자 정의 함수
def AutoFitColumnSize(worksheet, columns=None, margin=2):
    # 엑셀의 열 넓이를 자동으로 조정해주는 함수
    for i, column_cells in enumerate(worksheet.columns):
        is_ok = False
        if columns == None:
            is_ok = True
        elif isinstance(columns, list) and i in columns:
            is_ok = True
            
        if is_ok:
            length = max(len(str(cell.value)) for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = length + margin

    return worksheet

# 농업단순종사원, 50개 라는 조건을 포함한 워크넷 URL
url = 'https://www.work.go.kr/empInfo/empInfoSrch/list/dtlEmpSrchList.do?careerTo=&keywordJobCd=&occupation=905001&templateInfo=&shsyWorkSecd=&rot2WorkYn=&payGbn=&resultCnt=50&keywordJobCont=N&cert=&cloDateStdt=&moreCon=&minPay=&codeDepth2Info=11000&isChkLocCall=&sortFieldInfo=DATE&major=&resrDutyExcYn=&eodwYn=&sortField=DATE&staArea=&sortOrderBy=DESC&keyword=&termSearchGbn=all&carrEssYns=&benefitSrchAndOr=O&disableEmpHopeGbn=&webIsOut=&actServExcYn=&maxPay=&keywordStaAreaNm=N&emailApplyYn=&listCookieInfo=DTL&pageCode=&codeDepth1Info=11000&keywordEtcYn=&publDutyExcYn=&keywordJobCdSeqNo=&exJobsCd=&templateDepthNmInfo=&computerPreferential=&regDateStdt=&employGbn=&empTpGbcd=1&region=&infaYn=&resultCntInfo=50&siteClcd=all&cloDateEndt=&sortOrderByInfo=DESC&currntPageNo=1&indArea=&careerTypes=&searchOn=Y&tlmgYn=&subEmpHopeYn=&academicGbn=&templateDepthNoInfo=&foriegn=&mealOfferClcd=&station=&moerButtonYn=&holidayGbn=&srcKeyword=&enterPriseGbn=all&academicGbnoEdu=noEdu&cloTermSearchGbn=all&keywordWantedTitle=N&stationNm=&benefitGbn=&keywordFlag=&notSrcKeyword=&essCertChk=&isEmptyHeader=&depth2SelCode=&_csrf=494a8097-58f4-4b34-8c62-38c9a9e2c5d6&keywordBusiNm=N&preferentialGbn=&rot3WorkYn=&pfMatterPreferential=&regDateEndt=&staAreaLineInfo1=11000&staAreaLineInfo2=1&pageIndex=1&termContractMmcnt=&careerFrom=&laborHrShortYn=#viewSPL'

# 엑셀파일 생성 
wb=Workbook()
ws=wb.active
ws.title='워크넷 농업단순종사원'
ws.append(['', '회사명/정보 제공처', '채용공고명/담당업무/지원자격', '근무조건', '등록/마감일', '링크'])

# 크롬웹드라이버를 통해 셀레니움 사용
driver = webdriver.Chrome('./chromdriver')
driver.get(url)
html = driver.page_source
soup = BeautifulSoup(html, 'html.parser')

# 구인구직은 해당 사이트의 6번째 tbody(표) 부분에 있으므로 6번째 tbody에서 tr -> td 를 찾아 각 셀을 가져옴
list=soup.select('tbody')
list=list[5].select('tr > td')

# 표는 50행 5열로 총 250개의 셀을 가지고 있으며 /5를 통해 각 행마다 작업을 수행
# 1열은 기업비교 기능을 위한 체크박스 list[5i]
# 2열은 회사명 혹은 정보 제공처 list[5i+1]
# 3열은 채용공고명, 담당업무, 지원자격에 대한 요약 list[5i+2]
# 4열은 근무조건 list[5i+3]
# 5열은 등록일, 마감일에 대한 정보를 가지고 있음 list[5i+4]
for i in range(int(len(list)/5)):
    # 
    # 2열에서 정보를 가져오면 되나 a태그로 되어있는 기업도 있고, 단순 텍스트로 되어있는 경우도 있음
    # 따라서 1열에서 기업정보를 가져옴
    # <input type="checkbox" id="~~~~" value="~~~|VALIDATION|companyName|Title" title="~~~">
    # input 태그를 찾으면 다음과 같은 내용이 나오는데 여기서 | 를기준으로 문자열을 잘라 회사이름을 추출
    company=str(list[5*i].find('input'))
    company=company.split('|')
    name = company[2]

    # 채용공고명와 담당업무, 지원자격등을 추출
    # 채용공고명은 html 크롤링으로 가져올 수 있지만 담당업무, 지원자격은 js라서 웹드라이버 셀레니움을 사용
    text=str(list[5*i+2].find('div',class_='cp-info-in').get_text( ))
    text=text.strip()
    text = text+ ' ' + driver.find_element(By.ID, 'jobContLine'+str(i+1)).text
    
    # 근무 조건을 추출
    working_conditions = list[5*i+3].select('div > p')
    working_conditions = ('\n'.join(working_conditions[0].get_text().split()) + ' ' +'\n'.join(working_conditions[1].get_text().split())  
                        + ' ' +'\n'.join(working_conditions[2].get_text().split()) + ' ' +'\n'.join(working_conditions[3].get_text().split()))

    # 공고가 올라온 날짜, 마감일시를 가져옴
    date = list[5*i+4].select('div > p')
    date = '\n'.join(date[0].get_text().split()) + ' ' +'\n'.join(date[1].get_text().split())

    # 공고와 연결된 링크를 가져옴
    link='https://www.work.go.kr' + list[5*i+2].select_one('div > div > a')['href']
    
    data=[i+1, name, text, working_conditions, date, link]
    ws.append(data)

# 공고의 링크가 단순한 텍스트 이므로 하이퍼링크로 변환하는 과정
for i in range(2, ws.max_row +1):
    ws["F" + str(i)].hyperlink = ws["F" + str(i)].value
    ws['F'+str(i)].value = 'Link'
    ws["F" + str(i)].style = "Hyperlink"

# 엑셀을 보기좋게 디자인 변경하는 과정
AutoFitColumnSize(ws, margin=5)
ws['A1'].border = Border(top=Side(border_style='medium', color='000000'),
                         bottom=Side(border_style='medium', color='000000'))
ws['A1'].font = Font(bold=True)
ws['B1'].border = Border(top=Side(border_style='medium', color='000000'),
                         bottom=Side(border_style='medium', color='000000'))
ws['B1'].font = Font(bold=True)
ws['C1'].border =  Border(top=Side(border_style='medium', color='000000'),
                         bottom=Side(border_style='medium', color='000000'))
ws['C1'].font = Font(bold=True)
ws['D1'].border =  Border(top=Side(border_style='medium', color='000000'),
                         bottom=Side(border_style='medium', color='000000'))
ws['D1'].font = Font(bold=True)
ws['E1'].border =  Border(top=Side(border_style='medium', color='000000'),
                         bottom=Side(border_style='medium', color='000000'))
ws['E1'].font = Font(bold=True)
ws['F1'].border = Border(top=Side(border_style='medium', color='000000'),
                         right=Side(border_style='medium', color='000000'),
                         bottom=Side(border_style='medium', color='000000'))
ws['F1'].font = Font(bold=True)

# 엑셀로 저장
wb.save(str(datetime.datetime.now().date())+'.xlsx')
# 셀레니움 종료
driver.quit()